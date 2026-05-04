"""Tests for deterministic Excel compliance checks."""

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook

from backend.app.compliance.excel_checker import (
    ProductValue,
    check_excel_workbook,
    check_product_values,
    load_product_values,
)
from backend.main import app


def build_workbook(rows: list[tuple[object, ...]]) -> bytes:
    """Create an in-memory workbook for tests."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["parameter", "value", "unit", "category"])
    for row in rows:
        sheet.append(row)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_check_product_values_converts_protein_from_g_per_l() -> None:
    rows = [
        ProductValue("Energy", 2720, "kJ/L", "", ""),
        ProductValue("Protein", 15, "g/L", "milk-based", ""),
    ]

    result = check_product_values(rows)

    protein_result = result["results"][1]
    assert protein_result["status"] == "PASS"
    assert protein_result["converted_unit"] == "g/100 kJ"
    assert protein_result["converted_value"] == 0.551471
    assert "Converted from g/L" in protein_result["notes"]


def test_check_product_values_flags_failures_and_review_items() -> None:
    rows = [
        ProductValue("Docosahexaenoic acid", 13, "mg/100 kJ", "", ""),
        ProductValue("Protein", 12, "g/L", "", ""),
        ProductValue("Unknown nutrient", 1, "g/100 kJ", "", ""),
    ]

    result = check_product_values(rows)

    assert result["summary"] == {"total": 3, "passed": 0, "failed": 1, "needs_review": 2}
    assert result["results"][0]["status"] == "FAIL"
    assert result["results"][1]["status"] == "NEEDS_REVIEW"
    assert result["results"][2]["status"] == "NEEDS_REVIEW"


def test_load_product_values_scans_rows_without_fixed_headers() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Nutrition Panel"
    sheet.append(["Product nutrition data"])
    sheet.append(["Energy", 2720, "kJ/L"])
    sheet.append(["Protein", 15, "g/L", "milk-based"])
    output = BytesIO()
    workbook.save(output)

    values = load_product_values(output.getvalue())

    assert values[0].parameter == "Energy"
    assert values[0].value == 2720
    assert values[0].unit == "kJ/L"
    assert values[0].category == "Nutrition Panel"
    assert values[1].parameter == "Protein"
    assert values[1].category == "milk-based"


def test_load_product_values_scans_multiple_sheets() -> None:
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = "Macronutrients"
    first_sheet.append(["Nutrient", "Amount", "Units"])
    first_sheet.append(["Energy", 2720, "kJ/L"])

    second_sheet = workbook.create_sheet("Fatty acids")
    second_sheet.append(["Total trans fatty acids", 3, "% of total fatty acids"])
    output = BytesIO()
    workbook.save(output)

    values = load_product_values(output.getvalue())

    assert [value.parameter for value in values] == ["Energy", "Total trans fatty acids"]
    assert values[1].category == "Fatty acids"


def test_load_product_values_uses_nutrient_column_as_parameter_anchor() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Step 1 Review"
    sheet.append(["Core Fields"])
    sheet.append(["Nutrient", "Unit", "Can Target (= Total x 0.976)", "Old FSANZ Min", "Status"])
    sheet.append(["Protein (Nx 6.25)", "g/100 kJ", 0.58, 0.43, "Pass"])
    sheet.append(["LA^", "%", 0.2299, 9, "Fail"])
    sheet.append(["DHA", "mg/100 kJ", 13, 12, "Fail"])
    output = BytesIO()
    workbook.save(output)

    values = load_product_values(output.getvalue())

    assert [value.parameter for value in values] == [
        "Protein (Nx 6.25)",
        "Protein (Nx 6.25)",
        "LA^",
        "LA^",
        "DHA",
        "DHA",
    ]
    assert all(value.parameter not in {"mg", "N/A", "Fail"} for value in values)
    assert values[0].unit == "g/100 kJ"
    assert values[0].category == "Can Target (= Total x 0.976)"


def test_load_product_values_reads_short_units_from_unit_column() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Step 1 Review"
    sheet.append(["Nutrient", "Unit", "Can Target", None])
    sheet.append(["Energy", "kJ", 2092.656, 2102])
    sheet.append(["Protein (Nx 6.25)", "g", 11.8096, 11.7])
    output = BytesIO()
    workbook.save(output)

    values = load_product_values(output.getvalue())

    assert [(value.parameter, value.unit, value.category) for value in values] == [
        ("Energy", "kJ", "Can Target"),
        ("Protein (Nx 6.25)", "g", "Can Target"),
    ]


def test_check_excel_endpoint_accepts_xlsx_upload() -> None:
    workbook_bytes = build_workbook(
        [
            ("Energy", 2720, "kJ/L", ""),
            ("Protein", 15, "g/L", "milk-based"),
        ]
    )

    client = TestClient(app)
    response = client.post(
        "/check-excel",
        files={
            "file": (
                "product_check.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["filename"] == "product_check.xlsx"
    assert body["summary"]["passed"] == 2
    assert body["results"][1]["source"] == "FSANZ Standard 2.9.1 section 2.9.1-6"


def test_check_excel_endpoint_rejects_non_xlsx_upload() -> None:
    client = TestClient(app)
    response = client.post(
        "/check-excel",
        files={"file": ("product_check.csv", b"parameter,value,unit\nEnergy,2720,kJ/L\n", "text/csv")},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Please upload a .xlsx file."


def test_ai_review_excel_endpoint_accepts_xlsx_upload(monkeypatch) -> None:
    workbook_bytes = build_workbook([("Vitamin A", 80, "ug RE/100 kJ", "infant formula")])

    def fake_ai_review_excel_workbook(content: bytes, focus_instructions: str = "") -> dict[str, object]:
        assert content == workbook_bytes
        assert focus_instructions == "Focus on vitamin A."
        return {
            "summary": {
                "total": 1,
                "passed": 1,
                "failed": 0,
                "needs_review": 0,
                "insufficient_context": 0,
            },
            "results": [
                {
                    "row_index": 1,
                    "parameter": "Vitamin A",
                    "input_value": 80.0,
                    "input_unit": "ug RE/100 kJ",
                    "category": "infant formula",
                    "status": "PASS",
                    "requirement": "Example requirement.",
                    "reasoning": "The value meets the requirement [Source 1].",
                    "citations": ["[Source 1]"],
                    "sources": [
                        {
                            "source_id": 1,
                            "filename": "FSANZ Schedule 29.pdf",
                            "chunk_index": 4,
                            "page_number": 12,
                            "distance": 0.1,
                            "similarity": 0.9,
                            "excerpt": "Example requirement.",
                        }
                    ],
                }
            ],
        }

    monkeypatch.setattr("backend.app.api.routes.ai_review_excel_workbook", fake_ai_review_excel_workbook)

    client = TestClient(app)
    response = client.post(
        "/review-excel-ai",
        data={"focus_instructions": "Focus on vitamin A."},
        files={
            "file": (
                "product_check.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["filename"] == "product_check.xlsx"
    assert body["summary"]["passed"] == 1
    assert body["results"][0]["citations"] == ["[Source 1]"]
