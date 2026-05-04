"""Excel parsing and deterministic checks for selected infant-formula rules.

This module scans uploaded workbooks into normalized product-value rows, then
applies explicit Python checks where rules have been coded. Broader AI-assisted
screening lives in `ai_excel_reviewer.py`.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


REQUIRED_COLUMNS = {"parameter", "value", "unit"}
HEADER_ALIASES = {
    "parameter": {"parameter", "nutrient", "component", "analyte", "item", "name", "description"},
    "value": {"value", "amount", "result", "quantity", "level", "target", "declared value"},
    "unit": {"unit", "units", "uom"},
    "category": {"category", "type", "product type", "class"},
    "notes": {"notes", "note", "comment", "comments"},
}
UNIT_TOKENS = (
    "kj",
    "kcal",
    "g",
    "mg",
    "ug",
    "µg",
    "mcg",
    "iu",
    "%",
    "re",
    "alpha-te",
    "l",
    "100",
)
MAX_INFERRED_ROWS = 200


@dataclass(frozen=True)
class ProductValue:
    """One row extracted from the uploaded workbook."""

    parameter: str
    value: float | None
    unit: str
    category: str
    notes: str


@dataclass(frozen=True)
class Rule:
    """One deterministic regulatory check."""

    key: str
    label: str
    min_value: float | None
    max_value: float | None
    unit: str
    source: str


RULES = {
    "energy": Rule(
        key="energy",
        label="Energy content",
        min_value=2510,
        max_value=2930,
        unit="kJ/L",
        source="FSANZ Standard 2.9.1 section 2.9.1-5",
    ),
    "protein_milk_based": Rule(
        key="protein_milk_based",
        label="Protein content, milk-based infant formula",
        min_value=0.43,
        max_value=0.72,
        unit="g/100 kJ",
        source="FSANZ Standard 2.9.1 section 2.9.1-6",
    ),
    "protein_non_milk_based": Rule(
        key="protein_non_milk_based",
        label="Protein content, non-milk-based infant formula",
        min_value=0.54,
        max_value=0.72,
        unit="g/100 kJ",
        source="FSANZ Standard 2.9.1 section 2.9.1-6",
    ),
    "dha": Rule(
        key="dha",
        label="Docosahexaenoic acid",
        min_value=None,
        max_value=12,
        unit="mg/100 kJ",
        source="FSANZ Schedule 29 section S29-4",
    ),
    "total_trans_fatty_acids": Rule(
        key="total_trans_fatty_acids",
        label="Total trans fatty acids",
        min_value=None,
        max_value=4,
        unit="% of total fatty acids",
        source="FSANZ Schedule 29 section S29-4",
    ),
}


def normalize_text(value: object) -> str:
    """Normalize text for loose matching."""
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def normalize_unit(value: object) -> str:
    """Normalize common unit spellings."""
    unit = normalize_text(value).replace(" ", "")
    aliases = {
        "kj/l": "kJ/L",
        "kilojoules/l": "kJ/L",
        "kilojoulesperlitre": "kJ/L",
        "g/100kj": "g/100 kJ",
        "gper100kj": "g/100 kJ",
        "g/l": "g/L",
        "gperlitre": "g/L",
        "mg/100kj": "mg/100 kJ",
        "mgper100kj": "mg/100 kJ",
        "%": "% of total fatty acids",
        "%totalfattyacids": "% of total fatty acids",
        "%oftotalfattyacids": "% of total fatty acids",
        "percentoftotalfattyacids": "% of total fatty acids",
    }
    return aliases.get(unit, str(value or "").strip())


def is_likely_unit(value: object) -> bool:
    """Return whether a text cell looks like a measurement unit."""
    text = normalize_text(value)
    if not text:
        return False
    compact = text.replace(" ", "")
    if compact in {
        "kj/l",
        "g/l",
        "g/100kj",
        "mg/100kj",
        "ug/100kj",
        "µg/100kj",
        "mcg/100kj",
        "%oftotalfattyacids",
        "%",
    }:
        return True
    return any(token in compact for token in UNIT_TOKENS) and any(char.isdigit() for char in compact)


def parse_float(value: object) -> float | None:
    """Parse numeric cell values while tolerating percentage strings."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("%", "")
    try:
        return float(text)
    except ValueError:
        return None


def map_header_cell(value: object) -> str | None:
    """Map a workbook header cell to a normalized internal column name."""
    text = normalize_text(value)
    if not text:
        return None
    for column, aliases in HEADER_ALIASES.items():
        if text in aliases:
            return column
    return None


def find_header_map(rows: list[tuple[object, ...]]) -> tuple[int, dict[str, int]] | None:
    """Find a flexible parameter/value/unit header row if one exists."""
    for row_index, row in enumerate(rows[:25]):
        header_map: dict[str, int] = {}
        for column_index, cell in enumerate(row):
            mapped = map_header_cell(cell)
            if mapped and mapped not in header_map:
                header_map[mapped] = column_index
        if REQUIRED_COLUMNS <= set(header_map):
            return row_index, header_map
    return None


def row_cell(row: tuple[object, ...], index: int | None) -> object:
    """Safely return a row cell."""
    if index is None or index >= len(row):
        return ""
    return row[index]


def values_from_header_table(
    rows: list[tuple[object, ...]],
    header_index: int,
    header_map: dict[str, int],
    sheet_name: str,
) -> list[ProductValue]:
    """Extract values from a recognized table with parameter/value/unit columns."""
    values: list[ProductValue] = []
    for row in rows[header_index + 1 :]:
        if not any(cell not in (None, "") for cell in row):
            continue

        parameter = str(row_cell(row, header_map["parameter"]) or "").strip()
        value = parse_float(row_cell(row, header_map["value"]))
        unit = str(row_cell(row, header_map["unit"]) or "").strip()
        category = str(row_cell(row, header_map.get("category")) or sheet_name).strip()
        notes = str(row_cell(row, header_map.get("notes")) or "").strip()

        if parameter and value is not None and unit:
            values.append(ProductValue(parameter, value, unit, category, notes))

    return values


def nearest_text_left(row: tuple[object, ...], numeric_index: int) -> str:
    """Find the nearest useful text label to the left of a numeric value."""
    for index in range(numeric_index - 1, -1, -1):
        cell = row[index]
        if isinstance(cell, str) and normalize_text(cell) and not is_likely_unit(cell):
            return cell.strip()
    return ""


def nearest_unit(row: tuple[object, ...], header_row: tuple[object, ...] | None, numeric_index: int) -> str:
    """Infer a unit from adjacent cells or a header row."""
    for index in range(numeric_index + 1, min(len(row), numeric_index + 4)):
        cell = row[index]
        if isinstance(cell, str) and is_likely_unit(cell):
            return cell.strip()

    for index in range(numeric_index - 1, max(-1, numeric_index - 4), -1):
        cell = row[index]
        if isinstance(cell, str) and is_likely_unit(cell):
            return cell.strip()

    if header_row and numeric_index < len(header_row):
        header = header_row[numeric_index]
        if isinstance(header, str) and is_likely_unit(header):
            return header.strip()
    return ""


def nearby_category(row: tuple[object, ...], numeric_index: int, sheet_name: str) -> str:
    """Infer optional category/type text near a value."""
    category_terms: list[str] = []
    for index in range(numeric_index + 1, min(len(row), numeric_index + 5)):
        cell = row[index]
        if not isinstance(cell, str):
            continue
        text = cell.strip()
        if not text or is_likely_unit(text):
            continue
        category_terms.append(text)

    if category_terms:
        return " ".join(category_terms)
    return sheet_name


def infer_values_from_sheet(rows: list[tuple[object, ...]], sheet_name: str) -> list[ProductValue]:
    """Infer product values from a sheet without requiring fixed headers."""
    values: list[ProductValue] = []
    recent_header: tuple[object, ...] | None = None

    for row_number, row in enumerate(rows, start=1):
        if not any(cell not in (None, "") for cell in row):
            continue

        text_cells = [cell for cell in row if isinstance(cell, str) and normalize_text(cell)]
        numeric_indexes = [index for index, cell in enumerate(row) if parse_float(cell) is not None]
        if text_cells and not numeric_indexes:
            recent_header = row
            continue

        for numeric_index in numeric_indexes:
            value = parse_float(row[numeric_index])
            if value is None:
                continue

            parameter = nearest_text_left(row, numeric_index)
            if not parameter and recent_header and numeric_index < len(recent_header):
                header = recent_header[numeric_index]
                if isinstance(header, str) and not is_likely_unit(header):
                    parameter = header.strip()

            unit = nearest_unit(row, recent_header, numeric_index)
            if parameter and unit:
                values.append(
                    ProductValue(
                        parameter=parameter,
                        value=value,
                        unit=unit,
                        category=nearby_category(row, numeric_index, sheet_name),
                        notes=f"Inferred from sheet {sheet_name}, row {row_number}.",
                    )
                )

            if len(values) >= MAX_INFERRED_ROWS:
                return values

    return values


def deduplicate_values(values: list[ProductValue]) -> list[ProductValue]:
    """Remove duplicate inferred rows while preserving order."""
    seen: set[tuple[str, float | None, str, str]] = set()
    unique_values: list[ProductValue] = []
    for value in values:
        key = (
            normalize_text(value.parameter),
            value.value,
            normalize_unit(value.unit),
            normalize_text(value.category),
        )
        if key not in seen:
            seen.add(key)
            unique_values.append(value)
    return unique_values


def load_product_values(workbook_bytes: bytes) -> list[ProductValue]:
    """Smart-scan all workbook sheets into normalized product values."""
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
    values: list[ProductValue] = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        header_result = find_header_map(rows)
        if header_result:
            header_index, header_map = header_result
            values.extend(values_from_header_table(rows, header_index, header_map, sheet.title))
        else:
            values.extend(infer_values_from_sheet(rows, sheet.title))

    values = deduplicate_values(values)
    if not values:
        raise ValueError(
            "Workbook scan did not find product values. Include parameter names, numeric values, "
            "and units near each other so the smart scanner can infer rows."
        )

    return values


def detect_rule_key(row: ProductValue) -> str | None:
    """Map an input row to one supported rule key."""
    parameter = normalize_text(row.parameter)
    category = normalize_text(row.category)
    combined = f"{parameter} {category}"

    if "energy" in parameter:
        return "energy"
    if "protein" in parameter:
        if "non milk" in combined or "non-milk" in combined or "soy" in combined:
            return "protein_non_milk_based"
        if "milk" in combined:
            return "protein_milk_based"
        return "protein_unspecified"
    if "docosahexaenoic" in combined or "dha" in combined:
        return "dha"
    if "trans" in combined and "fat" in combined:
        return "total_trans_fatty_acids"

    return None


def find_energy_kj_per_l(rows: list[ProductValue]) -> float | None:
    """Find energy content in kJ/L if provided."""
    for row in rows:
        if detect_rule_key(row) == "energy" and normalize_unit(row.unit) == "kJ/L":
            return row.value
    return None


def format_requirement(rule: Rule) -> str:
    """Return a readable requirement range."""
    if rule.min_value is not None and rule.max_value is not None:
        return f"{rule.min_value:g}-{rule.max_value:g} {rule.unit}"
    if rule.min_value is not None:
        return f">= {rule.min_value:g} {rule.unit}"
    if rule.max_value is not None:
        return f"<= {rule.max_value:g} {rule.unit}"
    return rule.unit


def compare_value(value: float, rule: Rule) -> str:
    """Return PASS or FAIL for a converted value."""
    if rule.min_value is not None and value < rule.min_value:
        return "FAIL"
    if rule.max_value is not None and value > rule.max_value:
        return "FAIL"
    return "PASS"


def convert_value(row: ProductValue, rule: Rule, energy_kj_per_l: float | None) -> tuple[float | None, str]:
    """Convert an input value to the rule unit when supported."""
    if row.value is None:
        return None, "Value is missing or not numeric."

    input_unit = normalize_unit(row.unit)
    if input_unit == rule.unit:
        return row.value, ""

    if rule.unit == "g/100 kJ" and input_unit == "g/L":
        if not energy_kj_per_l:
            return None, "Protein in g/L requires energy content in kJ/L for conversion."
        return row.value / energy_kj_per_l * 100, "Converted from g/L using energy content."

    return None, f"Unsupported unit conversion from {row.unit!r} to {rule.unit}."


def result_row(
    row: ProductValue,
    status: str,
    rule: Rule | None = None,
    converted_value: float | None = None,
    converted_unit: str = "",
    notes: str = "",
) -> dict[str, Any]:
    """Build a serializable compliance result row."""
    return {
        "parameter": row.parameter,
        "input_value": row.value,
        "input_unit": row.unit,
        "category": row.category,
        "converted_value": round(converted_value, 6) if converted_value is not None else None,
        "converted_unit": converted_unit,
        "requirement": format_requirement(rule) if rule else "",
        "status": status,
        "source": rule.source if rule else "",
        "notes": notes,
    }


def check_product_values(rows: list[ProductValue]) -> dict[str, Any]:
    """Check uploaded product values against the supported rule set."""
    energy_kj_per_l = find_energy_kj_per_l(rows)
    results: list[dict[str, Any]] = []

    for row in rows:
        rule_key = detect_rule_key(row)
        if rule_key is None:
            results.append(
                result_row(
                    row,
                    status="NEEDS_REVIEW",
                    notes="No supported deterministic rule is implemented for this parameter.",
                )
            )
            continue

        if rule_key == "protein_unspecified":
            results.append(
                result_row(
                    row,
                    status="NEEDS_REVIEW",
                    notes="Protein checks require category to state milk-based or non-milk-based.",
                )
            )
            continue

        rule = RULES[rule_key]
        converted_value, note = convert_value(row=row, rule=rule, energy_kj_per_l=energy_kj_per_l)
        if converted_value is None:
            results.append(result_row(row, status="NEEDS_REVIEW", rule=rule, notes=note))
            continue

        status = compare_value(converted_value, rule)
        results.append(
            result_row(
                row,
                status=status,
                rule=rule,
                converted_value=converted_value,
                converted_unit=rule.unit,
                notes=note,
            )
        )

    summary = {
        "total": len(results),
        "passed": sum(1 for row in results if row["status"] == "PASS"),
        "failed": sum(1 for row in results if row["status"] == "FAIL"),
        "needs_review": sum(1 for row in results if row["status"] == "NEEDS_REVIEW"),
    }

    return {"summary": summary, "results": results}


def check_excel_workbook(workbook_bytes: bytes) -> dict[str, Any]:
    """Parse and check an uploaded Excel workbook."""
    rows = load_product_values(workbook_bytes)
    return check_product_values(rows)
